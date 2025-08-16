import json
import os
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from time import sleep

import gspread
from gspread_formatting import CellFormat, Color, TextFormat, format_cell_range
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from driver_extension import DriverExtension


class RoomsCrawler(DriverExtension):
    START_INFO = []

    SAVE_JSONL: bool = False
    SAVE_EXCEL: bool = False
    SAVE_GSHEET: bool = True
    GOOGLE_CREDENTIALS_FILEPATH: str = "kosta_hristov_credentials.json"

    max_days: int = 1  # maximum number of days to click in some crawlers that require to click on each day

    NEXT_BUTTON_PRESSES_NUM: int = 5  # number of times to click next button to scrape more weeks/days to mark always disabled slots
    NEXT_DAYS_SELECTOR: str = ""
    NEXT_BUTTON_PRESS_WAIT_TIME: int = 7  # waiting time after next button presses

    def __init__(self, sitename):
        super().__init__()
        self.sitename = sitename

    def scrape_room(self):
        pass

    def get_disabled_slots(self):
        if not self.NEXT_DAYS_SELECTOR:
            return
        unique_entries = []
        no_new_counter = 0
        for _ in range(self.NEXT_BUTTON_PRESSES_NUM):
            self.scrape_room()

            has_new, unique_entries = self.has_new_unique_entries(unique_entries)
            if not has_new:
                no_new_counter += 1
                if no_new_counter > 1:
                    break
            else:
                no_new_counter = 0

            next_button = self.safe_find(self.NEXT_DAYS_SELECTOR, return_element=True)
            if next_button:
                self.make_click(next_button)
                self.sleep_random(self.NEXT_BUTTON_PRESS_WAIT_TIME * 0.7, self.NEXT_BUTTON_PRESS_WAIT_TIME * 1.3)

        self.entries = unique_entries

    def has_new_unique_entries(self, old_unique_entries):
        new_unique_entries = self.filter_unique_entries(self.entries)
        has_new = len(new_unique_entries) > len(old_unique_entries)
        return has_new, new_unique_entries

    def mark_always_disabled_slots(self):
        """
        Mark slots that are always disabled for a given weekday as disabled=True
        """

        total_counts = defaultdict(lambda: defaultdict(int))
        disabled_counts = defaultdict(lambda: defaultdict(int))

        # Count per weekday/time
        for entry in self.entries:
            date_obj = datetime.strptime(entry["date"], "%Y-%m-%d")
            weekday = date_obj.strftime("%A")
            time = entry["time"]

            total_counts[weekday][time] += 1
            if not entry.get("available"):
                disabled_counts[weekday][time] += 1

        # Mark always-disabled slots
        for weekday, times in disabled_counts.items():
            for t, d_count in times.items():
                if d_count > 1 and d_count == total_counts[weekday][t]:  # disabled every time on this weekday
                    for entry in self.entries:
                        date_obj = datetime.strptime(entry["date"], "%Y-%m-%d")
                        if date_obj.strftime("%A") == weekday and entry["time"] == t:
                            entry["disabled"] = True

    def save_gsheet(self):
        room = self.scraped_data["room"]
        shortcut = self.scraped_data["shortcut"]
        room_url = self.scraped_data.get("url", "")

        try:
            # Setup credentials
            if not self.GOOGLE_CREDENTIALS_FILEPATH:
                self.logger.warning("❌ GOOGLE_CREDENTIALS_FILEPATH not set.")
                return
            if not Path(self.GOOGLE_CREDENTIALS_FILEPATH).exists():
                self.logger.warning(f"❌ {self.GOOGLE_CREDENTIALS_FILEPATH} does not exist.")
                return
            with open(self.GOOGLE_CREDENTIALS_FILEPATH) as f:
                credentials = json.load(f)
                required_keys = ["client_email", "private_key"]
                if not all(key in credentials for key in required_keys):
                    self.logger.warning(f"❌ {self.GOOGLE_CREDENTIALS_FILEPATH} is missing required data.")
                    return

            scope = [
                "https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive",
            ]
            creds = ServiceAccountCredentials.from_json_keyfile_name(self.GOOGLE_CREDENTIALS_FILEPATH, scope)
            client = gspread.authorize(creds)

            # Access the sheet and worksheet
            sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", shortcut).group(1)
            gid_match = re.search(r"gid=(\d+)", shortcut)
            target_gid = gid_match.group(1) if gid_match else "0"

            sheet = client.open_by_key(sheet_id)
            worksheet = None
            for ws in sheet.worksheets():
                if str(ws._properties.get("sheetId")) == target_gid:
                    worksheet = ws
                    break
            if not worksheet:
                self.logger.warning(f"❌ Worksheet with gid={target_gid} not found.")
                return

            # Prepare data
            entries = self.scraped_data["scraped_data"]
            try:
                target_date = entries[0]["date"]
            except IndexError:
                return
            target_date_dt = datetime.strptime(target_date, "%Y-%m-%d").date()

            # === Multi-week layout (repeat your existing weekly block for every week present in entries) ===

            # Build availability lookup & times (kept exactly as your logic expects)
            lookup = {(e["time"], e["date"]): e for e in entries if e.get("time")}
            used_times = sorted(set(e["time"] for e in entries if e.get("time")))

            # Group scraped dates by their ISO week start (Monday)
            unique_dates = sorted({e["date"] for e in entries})
            weeks_map = {}  # { week_start_date(date): set of date_str in that week that we actually have data for }
            for d_str in unique_dates:
                dt = datetime.strptime(d_str, "%Y-%m-%d").date()
                ws_start = dt - timedelta(days=dt.weekday())
                weeks_map.setdefault(ws_start, set()).add(d_str)

            # Iterate weeks in chronological order
            for week_start in sorted(weeks_map.keys()):
                # Your week structures (identical format to original single-week block)
                week_dates = [week_start + timedelta(days=i) for i in range(7)]
                week_date_strs = [d.strftime("%Y-%m-%d") for d in week_dates]
                headers = [d.strftime("%d-%m %a") for d in week_dates]

                # Ensure headers exist (reuse your matching rule B..H == headers)
                all_values = worksheet.get_all_values()
                header_row = None
                for r, row in enumerate(all_values, start=1):
                    if row[1:8] == headers:
                        header_row = r
                        break

                if header_row is None:
                    header_row = len(all_values) + 3
                    for i, header in enumerate(headers):
                        col_letter = chr(ord("B") + i)
                        worksheet.update(f"{col_letter}{header_row}", [[header]])
                        sleep(1)
                        format_cell_range(worksheet, f"{col_letter}{header_row}", CellFormat(textFormat=TextFormat(bold=True)))
                        sleep(1)

                # Mark current date header in pink (only if today is in THIS week’s 7-day header)
                today_str = datetime.now().strftime("%Y-%m-%d")
                if today_str in week_date_strs:
                    today_col_idx = week_date_strs.index(today_str)
                    today_col_letter = chr(ord("B") + today_col_idx)
                    format_cell_range(
                        worksheet,
                        f"{today_col_letter}{header_row}",
                        CellFormat(backgroundColor=Color(1.0, 0.8, 0.9)),  # pink
                    )
                    sleep(1)

                # Add week start in full format at A-(header_row - 1) just like before
                week_start_full = week_start.strftime("%Y.%m.%d")
                worksheet.update(f"A{header_row - 1}", [[week_start_full]])
                sleep(1)
                format_cell_range(
                    worksheet, f"A{header_row - 1}", CellFormat(textFormat=TextFormat(italic=True, fontSize=10), horizontalAlignment="LEFT")
                )
                sleep(1)

                # Fill grid (dates across columns B..H, times down from header_row+1)
                # Only write columns for dates we actually have in entries for this week
                for date in week_date_strs:
                    if date not in weeks_map[week_start]:
                        continue  # skip this column if no data for that date

                    col_idx = week_date_strs.index(date)
                    col_letter = chr(ord("B") + col_idx)

                    for i, time in enumerate(used_times):
                        row_num = header_row + 1 + i

                        # Write time in column A if empty (same as your logic)
                        time_cell_value = worksheet.cell(row_num, 1).value
                        if not time_cell_value:
                            worksheet.update(f"A{row_num}", [[time]])
                            sleep(1)

                        entry = lookup.get((time, date))
                        if not entry:
                            continue

                        # Color formatting (unchanged)
                        if entry.get("disabled"):
                            fmt = CellFormat(backgroundColor=Color(1.0, 0.6, 0.6))  # red
                        elif entry.get("available") is True:
                            fmt = CellFormat(backgroundColor=Color(0.8, 1.0, 0.8))  # green
                        elif entry.get("available") is False:
                            fmt = CellFormat(backgroundColor=Color(1.0, 0.85, 0.6))  # orange
                        else:
                            fmt = None

                        # Write time string into the date column if empty (same as before)
                        cell_value = worksheet.cell(row_num, col_idx + 2).value
                        if not cell_value:
                            worksheet.update(f"{col_letter}{row_num}", [[time]])
                            sleep(1)

                        if fmt:
                            format_cell_range(worksheet, f"{col_letter}{row_num}", fmt)
                            sleep(1)

                # Optional: fill orange background for whole column if "no time & available=False" for that date
                for i, date in enumerate(week_date_strs):
                    if date not in weeks_map[week_start]:
                        continue
                    has_empty_time_false = any(
                        e["date"] == date and not e.get("time") and e.get("available") is False for e in entries
                    )
                    if has_empty_time_false:
                        col_letter = chr(ord("B") + i)
                        for j in range(header_row + 1, header_row + 1 + len(used_times)):
                            val = worksheet.cell(j, i + 2).value
                            if not val:
                                format_cell_range(worksheet, f"{col_letter}{j}", CellFormat(backgroundColor=Color(1.0, 0.85, 0.6)))
                                sleep(1)
            # === end multi-week block ===

            self.logger.info(f"✅ Sheet written successfully for {room}")

        except Exception as e:
            self.logger.exception(f"❌ Failed to write to sheet: {e}")

    def save_jsonl(self):
        room_name = self.scraped_data["room"]
        room_name = re.sub(r"[^\w\s.-]", "", room_name)
        filename = f"{room_name.lower().replace(' ', '_')}.jsonl"
        filename = os.path.join("results", filename)

        scraped_data = self.scraped_data.get("scraped_data", [])

        with open(filename, "a", encoding="utf-8") as file:
            json.dump(scraped_data, file, ensure_ascii=False)
            file.write("\n")

    def save_excel(self, output_path="Escape Rooms - Statuses.xlsx"):
        entries = self.scraped_data["scraped_data"]
        room = self.scraped_data["room"]
        room_url = self.scraped_data.get("url", "")

        output_path = os.path.join("results", output_path)

        if Path(output_path).exists():
            wb = load_workbook(output_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        if room[:31] in wb.sheetnames:
            ws = wb[room[:31]]
        else:
            ws = wb.create_sheet(title=room[:31])
            ws.column_dimensions["A"].width = 30
            # Set weekday headers
            for i in range(7):
                ws.column_dimensions[get_column_letter(i + 2)].width = 13
            ws.cell(row=1, column=3, value=room_url)  # URL in C1

            ws.cell(row=1, column=1, value=room)  # Room name in A1

        if len(entries) == 0:
            return
        # Get the date from data (only 1 day expected)
        target_date = entries[0]["date"]
        target_date_dt = datetime.strptime(target_date, "%Y-%m-%d").date()

        # Determine the Monday of the week
        week_start = target_date_dt - timedelta(days=target_date_dt.weekday())
        week_dates = [(week_start + timedelta(days=i)) for i in range(7)]
        week_date_strs = [d.strftime("%Y-%m-%d") for d in week_dates]
        headers = [d.strftime("%d-%m %a") for d in week_dates]

        # Check if this week's header already exists
        week_found = False
        for row in range(1, ws.max_row + 1):
            values = [ws.cell(row=row, column=col).value for col in range(2, 9)]
            if values == headers:
                header_row = row
                week_found = True
                break

        if not week_found:
            # Find first empty row block
            row = ws.max_row + 3
            header_row = row
            for i, header in enumerate(headers, start=2):
                cell = ws.cell(row=header_row, column=i)
                cell.value = header
                cell.font = Font(bold=True)

            # Write week start date in A{header_row - 1}
            week_start_full = week_start.strftime("%d.%m.%Y")
            date_cell = ws.cell(row=header_row - 1, column=1)
            date_cell.value = week_start_full
            date_cell.font = Font(italic=True, size=10)
            date_cell.alignment = Alignment(horizontal="left")

        # Build lookup
        lookup = {(e["time"], e["date"]): e["available"] for e in entries if e.get("time") and "available" in e}
        used_times = sorted(set(e["time"] for e in entries if e.get("time")))

        # Write time slots vertically
        target_col = week_date_strs.index(target_date) + 2  # column B=2
        for i, time in enumerate(used_times):
            row = header_row + 1 + i
            cell = ws.cell(row=row, column=target_col)
            cell.value = time

            availability = lookup.get((time, target_date))
            if availability is True:
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # green
            elif availability is False:
                cell.fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")  # orange

        wb.save(output_path)
        self.logger.info(f"✅ Excel updated: {output_path}")

    def filter_unique_entries(self, entries):
        unique_entries = []
        seen = set()
        for entry in entries:
            # We'll create a unique key based on date, time, and availability
            key = (entry["date"], entry["time"], entry["available"])
            if key not in seen:
                seen.add(key)
                unique_entries.append(entry)

        return unique_entries

    def filter_scraped_data(self):
        entries = self.scraped_data["scraped_data"]

        unique_entries = []
        seen = set()
        for entry in entries:
            # We'll create a unique key based on date, time, and availability
            key = (entry["date"], entry["time"], entry["available"])
            if key not in seen:
                seen.add(key)
                unique_entries.append(entry)

        entries = unique_entries

        now = datetime.now()

        # Get sorted list of dates that have valid time + availability info
        valid_dates = sorted(set(e["date"] for e in entries if e.get("time") and isinstance(e.get("available"), bool)))

        # Filter out dates that are older than current day
        valid_dates = [d for d in valid_dates if datetime.strptime(d, "%Y-%m-%d").date() >= datetime.now().date()][:30]

        filtered = []
        for date_str in valid_dates:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

            for e in entries:
                if e.get("date") != date_str:
                    continue
                if not e.get("time") or not isinstance(e.get("available"), bool):
                    continue
                try:
                    recorded_time = e["time"].split("-")[0].strip()
                    time_obj = datetime.strptime(recorded_time, "%H:%M").time()
                except ValueError:
                    if e["time"] != "full day":
                        continue
                if date_obj == now.date():
                    if time_obj <= now.time():
                        continue
                filtered.append(e)

        if filtered:
            self.scraped_data["scraped_data"] = filtered

    def run(self):
        if self.USE_DRIVER:
            self.init_driver()
        try:
            for start_info in self.START_INFO:
                self.logger.info(f"Scraping room - {start_info['room']}")
                if self.USE_DRIVER:
                    self.load_html(start_info["url"], wait_time=10)
                self.entries = []
                self.scrape_room()
                self.get_disabled_slots()
                self.mark_always_disabled_slots()
                try:
                    self.entries.sort(key=lambda entry: (entry.get("date", ""), entry.get("time", "").split("-")[0].strip()))
                except Exception:
                    pass
                self.scraped_data = {
                    "room": start_info["room"],
                    "shortcut": start_info["shortcut"],
                    "url": start_info["url"],
                    "scrape_date": datetime.now().strftime("%Y-%m-%d"),
                    "scraped_data": self.entries,
                }
                self.filter_scraped_data()
                os.makedirs("results", exist_ok=True)
                if self.SAVE_JSONL:
                    self.save_jsonl()
                if self.SAVE_GSHEET:
                    self.save_gsheet()
                if self.SAVE_EXCEL:
                    self.save_excel()
        except Exception as e:
            self.logger.exception(f"Unexpected error: {e}")
        finally:
            if self.USE_DRIVER:
                self.quit_driver()
